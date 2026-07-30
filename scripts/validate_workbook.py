import argparse
import json
import math
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

REQUIRED_SHEETS = [
    "Config",
    "Branches",
    "Students",
    "Coaches",
    "Classes",
    "Sessions",
    "Enrollments",
    "Attendance",
    "Payments",
    "CoachPayments",
    "CoachAttendance",
]

ID_SHEETS = [
    "Branches",
    "Students",
    "Coaches",
    "Classes",
    "Sessions",
    "Enrollments",
    "Payments",
    "CoachPayments",
    "CoachAttendance",
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def key(value):
    value = clean(value)
    if value == "":
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip()


def date_value(value):
    value = clean(value)
    if value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        return text[:10]
    return text


def time_value(value):
    value = clean(value)
    if value == "":
        return ""
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def unique_headers(values):
    counts = Counter()
    headers = []
    for value in values:
        name = str(value or "").strip()
        counts[name] += 1
        headers.append(name if counts[name] == 1 else f"{name}__{counts[name]}")
    return headers


def load_sheet(workbook, name):
    worksheet = workbook[name]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = unique_headers(values[0])
    rows = []
    for row_number, values_row in enumerate(values[1:], start=2):
        if not any(clean(value) != "" for value in values_row):
            continue
        row = {headers[index]: clean(value) for index, value in enumerate(values_row)}
        row["_row"] = row_number
        rows.append(row)
    return rows


def add_error(report, sheet, row, code, detail):
    report["errors"].append({"sheet": sheet, "row": row, "code": code, "detail": detail})


def add_warning(report, sheet, row, code, detail):
    report["warnings"].append({"sheet": sheet, "row": row, "code": code, "detail": detail})


def require_reference(report, rows, sheet, column, target, target_ids):
    for row in rows.get(sheet, []):
        value = key(row.get(column))
        if value and value not in target_ids:
            add_error(report, sheet, row["_row"], "missing_reference", f"{column} does not exist in {target}")


def validate_unique_pair(report, rows, sheet, first, second):
    seen = {}
    for row in rows.get(sheet, []):
        pair = (key(row.get(first)), key(row.get(second)))
        if not all(pair):
            continue
        if pair in seen:
            add_error(report, sheet, row["_row"], "duplicate_relationship", f"Duplicates row {seen[pair]} for {first}+{second}")
        else:
            seen[pair] = row["_row"]


def validate_enum(report, rows, sheet, column, allowed, allow_blank=True):
    for row in rows.get(sheet, []):
        value = key(row.get(column))
        if not value and allow_blank:
            continue
        if value not in allowed:
            add_error(report, sheet, row["_row"], "invalid_enum", f"{column} has unsupported value {value!r}")


def validate(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    report = {
        "workbook": workbook_path.name,
        "sheets": {},
        "configKeys": [],
        "errors": [],
        "warnings": [],
        "ready": False,
    }
    if missing:
        for name in missing:
            add_error(report, name, 1, "missing_sheet", "Required worksheet is missing")
        return report

    rows = {name: load_sheet(workbook, name) for name in REQUIRED_SHEETS}
    report["sheets"] = {name: len(sheet_rows) for name, sheet_rows in rows.items()}
    report["configKeys"] = sorted(key(row.get("Key")) for row in rows["Config"] if key(row.get("Key")))

    ids = {}
    for sheet in ID_SHEETS:
        seen = {}
        ids[sheet] = set()
        for row in rows[sheet]:
            value = key(row.get("ID"))
            if not value:
                add_error(report, sheet, row["_row"], "missing_id", "ID is required")
                continue
            try:
                numeric = int(value)
                if numeric <= 0 or str(numeric) != value:
                    raise ValueError
            except ValueError:
                add_error(report, sheet, row["_row"], "invalid_id", "ID must be a positive integer")
                continue
            if value in seen:
                add_error(report, sheet, row["_row"], "duplicate_id", f"Duplicates row {seen[value]}")
            seen[value] = row["_row"]
            ids[sheet].add(value)

    require_reference(report, rows, "Students", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Coaches", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Classes", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Classes", "CoachID", "Coaches", ids["Coaches"])
    require_reference(report, rows, "Sessions", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Sessions", "ClassID", "Classes", ids["Classes"])
    require_reference(report, rows, "Sessions", "CoachID", "Coaches", ids["Coaches"])
    require_reference(report, rows, "Enrollments", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Enrollments", "StudentID", "Students", ids["Students"])
    require_reference(report, rows, "Enrollments", "ClassID", "Classes", ids["Classes"])
    require_reference(report, rows, "Attendance", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Attendance", "StudentID", "Students", ids["Students"])
    require_reference(report, rows, "Attendance", "SessionID", "Sessions", ids["Sessions"])
    require_reference(report, rows, "Attendance", "ClassID", "Classes", ids["Classes"])
    require_reference(report, rows, "Payments", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "Payments", "StudentID", "Students", ids["Students"])
    require_reference(report, rows, "Payments", "CoachID", "Coaches", ids["Coaches"])
    require_reference(report, rows, "Payments", "CoachPaymentID", "CoachPayments", ids["CoachPayments"])
    require_reference(report, rows, "CoachPayments", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "CoachPayments", "CoachID", "Coaches", ids["Coaches"])
    require_reference(report, rows, "CoachAttendance", "BranchID", "Branches", ids["Branches"])
    require_reference(report, rows, "CoachAttendance", "SessionID", "Sessions", ids["Sessions"])
    require_reference(report, rows, "CoachAttendance", "ClassID", "Classes", ids["Classes"])
    require_reference(report, rows, "CoachAttendance", "CoachID", "Coaches", ids["Coaches"])

    validate_unique_pair(report, rows, "Sessions", "ClassID", "Date")
    validate_unique_pair(report, rows, "Enrollments", "StudentID", "ClassID")
    validate_unique_pair(report, rows, "Attendance", "StudentID", "SessionID")
    validate_unique_pair(report, rows, "CoachAttendance", "SessionID", "CoachID")

    validate_enum(report, rows, "Branches", "Status", {"Active", "Inactive"}, False)
    validate_enum(report, rows, "Students", "Status", {"Active", "Trial", "Inactive"}, False)
    validate_enum(report, rows, "Students", "Gender", {"Male", "Female"})
    validate_enum(report, rows, "Students", "Level", {"Beginner", "Intermediate", "Advanced"})
    validate_enum(report, rows, "Classes", "DayOfWeek", {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}, False)
    validate_enum(report, rows, "Coaches", "Status", {"Active", "Inactive"}, False)
    validate_enum(report, rows, "Coaches", "CoachType", {"Head", "Assistant"})
    validate_enum(report, rows, "Attendance", "Status", {"Present", "Absent"})
    validate_enum(report, rows, "Payments", "Status", {"Paid", "Partial", "Unpaid"}, False)
    validate_enum(report, rows, "Payments", "Method", {"Cash", "Bank Transfer", "Touch 'n Go eWallet", "Online", "Others"})

    branch_by_student = {key(row.get("ID")): key(row.get("BranchID")) for row in rows["Students"]}
    branch_by_coach = {key(row.get("ID")): key(row.get("BranchID")) for row in rows["Coaches"]}
    branch_by_class = {key(row.get("ID")): key(row.get("BranchID")) for row in rows["Classes"]}
    class_by_session = {key(row.get("ID")): key(row.get("ClassID")) for row in rows["Sessions"]}
    branch_by_session = {key(row.get("ID")): key(row.get("BranchID")) for row in rows["Sessions"]}

    for row in rows["Classes"]:
        coach_id = key(row.get("CoachID"))
        if coach_id and branch_by_coach.get(coach_id) != key(row.get("BranchID")):
            add_error(report, "Classes", row["_row"], "cross_branch", "Coach and class branches differ")
    for row in rows["Sessions"]:
        class_id = key(row.get("ClassID"))
        coach_id = key(row.get("CoachID"))
        if class_id and branch_by_class.get(class_id) != key(row.get("BranchID")):
            add_error(report, "Sessions", row["_row"], "cross_branch", "Class and session branches differ")
        if coach_id and branch_by_coach.get(coach_id) != key(row.get("BranchID")):
            add_error(report, "Sessions", row["_row"], "cross_branch", "Coach and session branches differ")
    for row in rows["Enrollments"]:
        student_id = key(row.get("StudentID"))
        class_id = key(row.get("ClassID"))
        branches = {key(row.get("BranchID")), branch_by_student.get(student_id), branch_by_class.get(class_id)} - {None, ""}
        if len(branches) > 1:
            add_error(report, "Enrollments", row["_row"], "cross_branch", "Student, class, and enrollment branches differ")
    for row in rows["Attendance"]:
        student_id = key(row.get("StudentID"))
        session_id = key(row.get("SessionID"))
        class_id = key(row.get("ClassID"))
        if class_by_session.get(session_id) and class_by_session.get(session_id) != class_id:
            add_error(report, "Attendance", row["_row"], "class_mismatch", "Attendance ClassID differs from its session")
        branches = {key(row.get("BranchID")), branch_by_student.get(student_id), branch_by_session.get(session_id)} - {None, ""}
        if len(branches) > 1:
            add_error(report, "Attendance", row["_row"], "cross_branch", "Student, session, and attendance branches differ")

    for sheet, date_columns in {
        "Students": ["DOB", "CreatedAt"],
        "Branches": ["CreatedAt"],
        "Sessions": ["Date"],
        "Enrollments": ["StartDate"],
        "Attendance": ["Date"],
        "Payments": ["FeeMonth", "DateReceived"],
        "CoachPayments": ["DatePaid"],
        "CoachAttendance": ["Date"],
    }.items():
        for row in rows[sheet]:
            for column in date_columns:
                value = row.get(column)
                if clean(value) != "" and not date_value(value):
                    add_warning(report, sheet, row["_row"], "date_parse", f"{column} could not be normalized")

    for row in rows["Classes"]:
        if not key(row.get("Label")):
            add_error(report, "Classes", row["_row"], "missing_required", "Label is required")
        time_value(row.get("StartTime"))
        time_value(row.get("EndTime"))
    for row in rows["Students"]:
        if not key(row.get("Name")):
            add_error(report, "Students", row["_row"], "missing_required", "Name is required")
    for row in rows["Coaches"]:
        if not key(row.get("Name")):
            add_error(report, "Coaches", row["_row"], "missing_required", "Name is required")

    if not rows["Payments"]:
        add_warning(report, "Payments", 1, "empty_sheet", "No payment history will be imported")
    if not rows["CoachPayments"]:
        add_warning(report, "CoachPayments", 1, "empty_sheet", "No coach payout history will be imported")

    report["ready"] = len(report["errors"]) == 0
    return report


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    report = validate(args.workbook)
    if args.json:
        print(json.dumps(report, indent=2, default=str))
    else:
        print(f"Workbook: {report['workbook']}")
        for name, count in report["sheets"].items():
            print(f"  {name}: {count}")
        print(f"Errors: {len(report['errors'])}")
        print(f"Warnings: {len(report['warnings'])}")
        for issue in report["errors"] + report["warnings"]:
            print(f"  {issue['sheet']} row {issue['row']}: {issue['code']} - {issue['detail']}")
        print(f"Ready for transformation: {'yes' if report['ready'] else 'no'}")
    raise SystemExit(0 if report["ready"] else 1)


if __name__ == "__main__":
    main()
